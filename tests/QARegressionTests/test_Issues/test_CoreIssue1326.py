import os
import subprocess
import unittest
import openpyxl
import pytest
from conftest import get_python_executable
from QARegressionTests.globals import (
    dataset_details_sheet,
    issue_datails_sheet,
    rules_report_sheet,
    issue_sheet_record_column,
    issue_sheet_variable_column,
    issue_sheet_values_column,
)


# @pytest.mark.regression
class TestPrefTerm(unittest.TestCase):
    def test_positive_dataset(self):
        # Run the command in the terminal
        command = [
            f"{get_python_executable()}",
            "-m",
            "core",
            "validate",
            "-s",
            "usdm",
            "-v",
            "3-0",
            "-dp",
            os.path.join("tests", "resources", "CoreIssue1326", "regression-test-coreid-DDF00015-positive.json"),
            "-lr",
            os.path.join("tests", "resources", "CoreIssue1326", "rule.yml")
        ]
        subprocess.run(command, check=True)

        # Get the latest created Excel file
        files = os.listdir()
        excel_files = [
            file
            for file in files
            if file.startswith("CORE-Report-") and file.endswith(".xlsx")
        ]
        excel_file_path = sorted(excel_files)[-1]
        # # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file_path)

        # Go to the "Issue Details" sheet
        sheet = workbook[issue_datails_sheet]

        record_column = sheet[issue_sheet_record_column]
        variables_column = sheet[issue_sheet_variable_column]
        values_column = sheet[issue_sheet_values_column]

        record_values = [cell.value for cell in record_column[1:]]
        variables_values = [cell.value for cell in variables_column[1:]]
        values_column_values = [cell.value for cell in values_column[1:]]

        # Remove None values using list comprehension
        record_values = [value for value in record_values if value is not None]
        variables_values = [value for value in variables_values if value is not None]
        values_column_values = [
            value for value in values_column_values if value is not None
        ]
        rules_report_sheet = workbook["Rules Report"]
        rules_values = [row for row in rules_report_sheet.iter_rows(values_only=True)][1:]
        rules_values = [row for row in rules_values if any(row)]
        # Perform the assertion
        # Ensure only two negative values are caught
        assert rules_values[0][0] == "CORE-000409"
        assert rules_values[0][-1] == "SUCCESS"
        assert len(record_values) == 0
        assert len(variables_values) == 0
        assert len(values_column_values) == 0
        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)

    def test_negaive_dataset(self):
        # Run the command in the terminal
        command = [
            f"{get_python_executable()}",
            "-m",
            "core",
            "validate",
            "-s",
            "usdm",
            "-v",
            "3-0",
            "-dp",
            os.path.join("tests", "resources", "CoreIssue1326", "regression-test-coreid-DDF00015-negative.json"),
            "-lr",
            os.path.join("tests", "resources", "CoreIssue1326", "rule.yml")
        ]
        subprocess.run(command, check=True)

        # Get the latest created Excel file
        files = os.listdir()
        excel_files = [
            file
            for file in files
            if file.startswith("CORE-Report-") and file.endswith(".xlsx")
        ]
        excel_file_path = sorted(excel_files)[-1]
        # # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file_path)

        # --- Dataset Details ---
        dataset_sheet = workbook[dataset_details_sheet]
        dataset_values = [row for row in dataset_sheet.iter_rows(values_only=True)][1:]
        dataset_values = [row for row in dataset_values if any(row)]
        assert len(dataset_values) > 0
        assert dataset_values[0][0] == "StudyVersion.xpt"
        assert dataset_values[0][-1] == 7

        # --- Issue Summary ---
        issue_summary_sheet = workbook["Issue Summary"]
        summary_values = [row for row in issue_summary_sheet.iter_rows(values_only=True)][1:]
        summary_values = [row for row in summary_values if any(row)]
        assert len(summary_values) > 0
        assert summary_values[0][1] == "CORE-000409"
        assert summary_values[0][3] == 7

        # --- Issue Details ---
        expected_values = [
            "C199989, C199989, Phase II Trial, PHASE II TRIAL, C15601, Phase Ib Trial",
            "null, null, Not Applicable, NOT APPLICABLE, C48660, N/A",
            "C198366, C198366, null, null, C198366xx, Phase I/II/III Trial",
            "C15600, C15600, null, null, C00001x, Phase I Trial",
            "C198366, C198366, Phase II Trial, PHASE II TRIAL, C15601, Phase I/II/III Trial",
            "C199989, C199989, null, null, C198366xx, PHASE IB TRIAL",
            "null, null, Phase Ib Trial, PHASE IB TRIAL, C199989, PHASE 2 TRIAL",
        ]
        issue_details_sheet = workbook[issue_datails_sheet]
        details_values = [row for row in issue_details_sheet.iter_rows(values_only=True)][1:]
        details_values = [row for row in details_values if any(row)]
        assert all(row[0] == "CORE-000409" for row in details_values)
        actual_values = [row[-1] for row in details_values]
        assert actual_values == expected_values

        # --- Rules Report ---
        rules_report_sheet = workbook["Rules Report"]
        rules_values = [row for row in rules_report_sheet.iter_rows(values_only=True)][1:]
        rules_values = [row for row in rules_values if any(row)]
        assert len(rules_values) > 0
        assert rules_values[0][0] == "CORE-000409"
        assert rules_values[0][-1] == "SUCCESS"

        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)

# if __name__ == "__main__":
#     unittest.main()
