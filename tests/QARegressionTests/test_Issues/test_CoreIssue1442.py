import os
import subprocess
import unittest
import openpyxl
import pytest
from conftest import get_python_executable


@pytest.mark.regression
class TestCoreIssue1442(unittest.TestCase):
    def test_positive_dataset(self):
        # Run the command in the terminal
        command = [
            f"{get_python_executable()}",
            "-m",
            "core",
            "validate",
            "-s",
            "usdm",
            "-v",
            "4-0",
            "-d",
            os.path.join("tests", "resources", "CoreIssue1442"),
            "-ft",
            "json",
            "-lr",
            os.path.join("tests", "resources", "CoreIssue1442", "rule.yml"),
        ]
        subprocess.run(command, check=True)

        # Get the latest created Excel file
        files = os.listdir()
        excel_files = [
            file
            for file in files
            if file.startswith("CORE-Report-") and file.endswith(".xlsx")
        ]
        excel_file_path = sorted(excel_files)[-1]
        # # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file_path)
        if "Conformance Details" in workbook.sheetnames:
            conformance_sheet = workbook["Conformance Details"]
            found = False
            for row in conformance_sheet.iter_rows(min_row=2, values_only=True):
                for idx, cell in enumerate(row[:-1]):
                    if (
                        cell == "JSON file name"
                        and row[idx + 1] == "CDISC_Pilot_Study.json"
                    ):
                        found = True
                        break
                if found:
                    break
            assert (
                found
            ), "Pair ('JSON file name', 'CDISC_Pilot_Study.json') not found in any row of 'Conformance Details' sheet."
        else:
            assert False, "'Conformance Details' sheet not found in report."

        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)
