import os
import subprocess
import unittest
import openpyxl
import pytest
from conftest import get_python_executable


@pytest.mark.regression
class TestCoreIssue1487(unittest.TestCase):
    def test_positive_dataset(self):
        # Run the command in the terminal
        command = [
            f"{get_python_executable()}",
            "-m",
            "core",
            "validate",
            "-s",
            "sdtmig",
            "-v",
            "5-0",
            "-d",
            os.path.join("tests", "resources", "CoreIssue1487"),
            "-r",
            "CORE-000354",
        ]
        subprocess.run(command, check=True)

        # Get the latest created Excel file
        files = os.listdir()
        excel_files = [
            file
            for file in files
            if file.startswith("CORE-Report-") and file.endswith(".xlsx")
        ]
        excel_file_path = sorted(excel_files)[-1]
        # # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file_path)

        assert "Rules Report" in workbook.sheetnames
        rules_sheet = workbook["Rules Report"]
        target_row = None
        for row in rules_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == "CORE-000354":
                target_row = row
                break
        assert target_row, "Rule CORE-000354 not present in 'Rules Report' sheet."
        assert (
            target_row[4] and "was requested but is not available" in target_row[4]
        ), "Expected error message for CORE-000354 not found."
        assert target_row[5] == "SKIPPED", "CORE-000354 status should be SKIPPED."

        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)
