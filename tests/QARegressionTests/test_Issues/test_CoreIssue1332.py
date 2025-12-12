import os
import subprocess
import openpyxl
import pytest
from conftest import get_python_executable
from QARegressionTests.globals import (
    issue_datails_sheet,
    rules_report_sheet,
)


@pytest.mark.regression
def test_ap_domain_should_be_correctly_substituted():
    # Run the command in the terminal
    command = [
        f"{get_python_executable()}",
        "-m",
        "core",
        "validate",
        "-s",
        "sdtmig",
        "-v",
        "3-3",
        "-dp",
        os.path.join(
            "tests",
            "resources",
            "CoreIssue1332",
            "Datasets.json",
        ),
        "-lr",
        os.path.join("tests", "resources", "CoreIssue1332", "Rule.yml"),
    ]
    subprocess.run(command, check=True)

    files = os.listdir()
    excel_files = [
        file
        for file in files
        if file.startswith("CORE-Report-") and file.endswith(".xlsx")
    ]
    excel_file_path = sorted(excel_files)[-1]

    # # Open the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)

    # Go to the "Issue Details" sheet
    sheet = workbook[issue_datails_sheet]
    details_values = [row for row in sheet.iter_rows(values_only=True)][1:]
    details_values = [row for row in details_values if any(row)]
    assert len(details_values) == 8

    # CORE-ID
    assert all(row[0] == "CORE-000181" for row in details_values)

    # Message
    assert (
        len(
            [
                row
                for row in details_values
                if row[1] == "APMH Domain value length is not equal to 4."
            ]
        )
        == 4
    )
    assert (
        len(
            [
                row
                for row in details_values
                if row[1] == "APSQ Domain value length is not equal to 4."
            ]
        )
        == 4
    )

    # Variable(s)
    assert len([row for row in details_values if row[7] == "MHTERM"]) == 4
    assert len([row for row in details_values if row[7] == "SQTERM"]) == 4

    # Value(s)
    assert all(row[8] == "POMPE DISEASE" for row in details_values)

    # Go to the "Issue Summary" sheet
    issue_summary_sheet = workbook["Issue Summary"]
    summary_values = [row for row in issue_summary_sheet.iter_rows(values_only=True)][
        1:
    ]
    summary_values = [row for row in summary_values if any(row)]
    assert len(summary_values) == 2
    assert all(row[1] == "CORE-000181" for row in summary_values)  # CORE-ID
    assert (
        summary_values[0][2] == "APMH Domain value length is not equal to 4."
    )  # Message
    assert (
        summary_values[1][2] == "APSQ Domain value length is not equal to 4."
    )  # Message
    assert all(row[3] == 4 for row in summary_values)  # Issues

    # --- Rules Report ---
    rules_values = [
        row for row in workbook[rules_report_sheet].iter_rows(values_only=True)
    ][1:]
    rules_values = [row for row in rules_values if any(row)]
    assert len(rules_values) == 1
    assert rules_values[0][0] == "CORE-000181"
    assert "SUCCESS" in rules_values[0]

    if os.path.exists(excel_file_path):
        os.remove(excel_file_path)
