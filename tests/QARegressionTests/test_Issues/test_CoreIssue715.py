import os
import subprocess
import unittest
import openpyxl
import pytest
from conftest import get_python_executable
from QARegressionTests.globals import (
    issue_datails_sheet,
    rules_report_sheet,
    issue_sheet_record_column,
    issue_sheet_variable_column,
    issue_sheet_values_column,
)


@pytest.mark.regression
class TestCoreIssue715(unittest.TestCase):
    def test_positive_dataset(self):
        # Run the command in the terminal
        command = [
            f"{get_python_executable()}",
            "-m",
            "core",
            "validate",
            "-s",
            "usdm",
            "-v",
            "4-0",
            "-dp",
            os.path.join(
                "tests", "resources", "CoreIssue715", "CDISC_Pilot_Study.json"
            ),
            "-lr",
            os.path.join("tests", "resources", "CoreIssue715", "rule.yml"),
        ]
        subprocess.run(command, check=True)

        # Get the latest created Excel file
        files = os.listdir()
        excel_files = [
            file
            for file in files
            if file.startswith("CORE-Report-") and file.endswith(".xlsx")
        ]
        excel_file_path = sorted(excel_files)[-1]
        # # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file_path)

        # Go to the "Issue Details" sheet
        sheet = workbook[issue_datails_sheet]

        record_column = sheet[issue_sheet_record_column]
        variables_column = sheet[issue_sheet_variable_column]
        values_column = sheet[issue_sheet_values_column]

        record_values = [cell.value for cell in record_column[1:]]
        variables_values = [cell.value for cell in variables_column[1:]]
        values_column_values = [cell.value for cell in values_column[1:]]

        # Remove None values using list comprehension
        record_values = [value for value in record_values if value is not None]
        variables_values = [value for value in variables_values if value is not None]
        values_column_values = [
            value for value in values_column_values if value is not None
        ]
        rules_values = [
            row for row in workbook[rules_report_sheet].iter_rows(values_only=True)
        ][1:]
        rules_values = [row for row in rules_values if any(row)]
        # Perform the assertion
        assert rules_values[0][0] == "CORE-000409"
        assert "SUCCESS" in rules_values[0]
        assert len(record_values) == 0
        assert len(variables_values) == 0
        assert len(values_column_values) == 0
        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)

    def test_negative_dataset(self):
        # Run validation for invalid JSON
        command = [
            f"{get_python_executable()}",
            "-m",
            "core",
            "validate",
            "-s",
            "usdm",
            "-v",
            "4-0",
            "-dp",
            os.path.join(
                "tests", "resources", "CoreIssue715", "CDISC_Pilot_Study_Invalid.json"
            ),
            "-lr",
            os.path.join("tests", "resources", "CoreIssue715", "rule.yml"),
        ]
        subprocess.run(command, check=True)

        files = os.listdir()
        excel_files = [
            f for f in files if f.startswith("CORE-Report-") and f.endswith(".xlsx")
        ]
        excel_file_path = sorted(excel_files)[-1]
        workbook = openpyxl.load_workbook(excel_file_path)

        # Issue Summary basic checks
        issue_summary_sheet = workbook["Issue Summary"]
        summary_values = [r for r in issue_summary_sheet.iter_rows(values_only=True)][
            1:
        ]
        summary_values = [r for r in summary_values if any(r)]
        assert summary_values and summary_values[0][1] == "CORE-000409"
        assert summary_values[0][3] == 1

        # Issue Details strict checks: now expect one row per error
        issue_details_sheet = workbook[issue_datails_sheet]
        details_rows = [r for r in issue_details_sheet.iter_rows(values_only=True)][1:]
        details_rows = [r for r in details_rows if any(r)]
        # Expect exactly 1 row
        assert len(details_rows) == 1

        # Expected exact strings
        for row in details_rows:
            assert row[0] == "CORE-000409"
            assert (
                row[1]
                == "The narrative content dataset is expected to be USDM JSON compliant."
            )
            assert row[7] == (
                "_path, error_attribute, error_context, error_value, id, "
                "instanceType, json_path, message, validator, validator_value"
            )

        # Rules Report
        rules_rows = [
            r for r in workbook[rules_report_sheet].iter_rows(values_only=True)
        ][1:]
        rules_rows = [r for r in rules_rows if any(r)]
        assert rules_rows and rules_rows[0][0] == "CORE-000409"
        assert "SUCCESS" in rules_rows[0]
        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)
