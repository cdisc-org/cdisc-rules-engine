import os
import subprocess
import unittest
import openpyxl
import pytest
from conftest import get_python_executable


@pytest.mark.regression
class TerminalCommandTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # Run the command in the terminal
        command = [
            f"{get_python_executable()}",
            "-m",
            "core",
            "validate",
            "-s",
            "sdtmig",
            "-v",
            "3.4",
            "-dp",
            os.path.join("resources", "datasets", "ae.xpt"),
        ]
        subprocess.run(command, check=True)

        # Get the latest created Excel file
        files = os.listdir()
        excel_files = [
            file
            for file in files
            if file.startswith("CORE-Report-") and file.endswith(".xlsx")
        ]
        cls.excel_file_path = sorted(excel_files)[-1]

    def test_excel_file_contents(self):
        # Check if the Excel file is created
        self.assertTrue(
            os.path.exists(self.excel_file_path),
            f"Excel file '{self.excel_file_path}' is not created.",
        )

        # Open the Excel file
        workbook = openpyxl.load_workbook(self.excel_file_path)

        # Check if "Bundle Details" sheet does not exist
        self.assertNotIn(
            "Bundle Details",
            workbook.sheetnames,
            "Sheet 'Bundle Details' should not exist.",
        )

        # Check if "Conformance Details" sheet exists
        self.assertIn(
            "Conformance Details",
            workbook.sheetnames,
            "Sheet 'Conformance Details' does not exist.",
        )

        # Check if "RULE-ID" column does not exist in:
        # Issue Summary, Issue Details, and Rules Report sheets
        for sheet_name in ["Issue Summary", "Issue Details", "Rules Report"]:
            sheet = workbook[sheet_name]
            column_names = [cell.value for cell in sheet[1]]
            self.assertNotIn(
                "RULE-ID",
                column_names,
                f"'RULE-ID' column should not exist in sheet '{sheet_name}'.",
            )

            self.assertIn(
                "CORE-ID",
                column_names,
                f"'CORE-ID' column should exist in sheet '{sheet_name}'.",
            )

    @classmethod
    def tearDownClass(cls):
        # Delete the Excel file
        if os.path.exists(cls.excel_file_path):
            os.remove(cls.excel_file_path)


if __name__ == "__main__":
    unittest.main()
