import os
import subprocess
import openpyxl
import pytest
from conftest import get_python_executable
from QARegressionTests.globals import (
    dataset_details_sheet,
    issue_datails_sheet,
    rules_report_sheet,
    issue_sheet_coreid_column,
)


@pytest.mark.regression
def test_validate_define_xml_against_lib_metadata():
    command = [
        f"{get_python_executable()}",
        "-m",
        "core",
        "validate",
        "-s",
        "sdtmig",
        "-v",
        "3-4",
        "-dp",
        os.path.join(
            "tests",
            "resources",
            "CoreIssue1421",
            "Dataset.json",
        ),
        "-lr",
        os.path.join("tests", "resources", "CoreIssue1421", "Rule.yml"),
        "-dxp",
        os.path.join("tests", "resources", "CoreIssue1421", "Define.xml"),
    ]
    subprocess.run(command, check=True)

    # Get the latest created Excel file
    files = os.listdir()
    excel_files = [
        file
        for file in files
        if file.startswith("CORE-Report-") and file.endswith(".xlsx")
    ]
    excel_file_path = sorted(excel_files)[-1]
    # Open the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)

    # Go to the "Issue Details" sheet
    sheet = workbook[issue_datails_sheet]

    expected_output_variables = [
        "variable_name",
        "library_variable_name",
        "library_variable_ccode",
        "library_variable_data_type",
        "define_variable_name",
        "define_variable_ccode",
    ]

    # Check Variable(s) column (H)
    variables_names_column = sheet["H"]
    variables_names_values = [
        cell.value for cell in variables_names_column[1:] if cell.value is not None
    ]

    # Collapsed behavior: one issue row that lists the reported output fields
    assert len(variables_names_values) == 1
    assert variables_names_values[0] == ", ".join(expected_output_variables)

    issue_rows = [
        row for row in sheet.iter_rows(min_row=2, values_only=True) if any(row)
    ]

    reported_issue_rows = [row for row in issue_rows if row[7]]
    execution_error_rows = [row for row in issue_rows if not row[7] and row[8]]

    assert len(reported_issue_rows) == 1
    assert len(execution_error_rows) == 1

    issue_row = reported_issue_rows[0]
    assert issue_row[7] == ", ".join(expected_output_variables)

    reported_values = [value.strip() for value in issue_row[8].split(",")]
    assert len(reported_values) == len(expected_output_variables)

    # Still verify the row is diagnostically useful
    assert reported_values[0] == reported_values[1]
    assert reported_values[0] == reported_values[4]
    assert reported_values[2] not in {"", "null"}
    assert reported_values[3] not in {"", "null"}
    assert reported_values[5] != reported_values[2]

    execution_error_row = execution_error_rows[0]
    assert execution_error_row[3] == "SUPPEC"
    assert "Failed to build dataset for rule validation" in execution_error_row[8]

    dataset_column = sheet["D"]
    dataset_column_values = [
        cell.value for cell in dataset_column[1:] if cell.value is not None
    ]
    assert sorted(set(dataset_column_values)) == ["DM", "SUPPEC"]

    core_id_column = sheet[issue_sheet_coreid_column]
    core_id_column_values = [
        cell.value for cell in core_id_column[1:] if cell.value is not None
    ]
    assert set(core_id_column_values) == {"CDISC.SDTMIG.CG0999"}

    # Go to the "Rules Report" sheet
    rules_values = [
        row for row in workbook[rules_report_sheet].iter_rows(values_only=True)
    ][1:]
    rules_values = [row for row in rules_values if any(row)]

    assert len(rules_values) == 1
    rule_row = rules_values[0]

    assert rule_row[0] == "CDISC.SDTMIG.CG0999"
    assert rule_row[4] == "Issue with codelist definition in the Define-XML document."
    assert rule_row[5] == "EXECUTION ERROR"

    # Go to the "Dataset Details" sheet
    dataset_sheet = workbook[dataset_details_sheet]
    dataset_values = [row for row in dataset_sheet.iter_rows(values_only=True)][1:]
    dataset_values = [row for row in dataset_values if any(row)]
    assert len(dataset_values) > 0
    dataset_names = set(row[0] for row in dataset_values if row[0] is not None)
    assert dataset_names == {"AE", "DM", "EC", "EX", "SUPPEC"}
    expected_records = {
        "AE": 74,
        "DM": 18,
        "EC": 1590,
        "EX": 1583,
        "SUPPEC": 13,
    }
    for row in dataset_values:
        dataset_name = row[0]
        records_count = row[-1]
        assert records_count == expected_records[dataset_name]

    # Go to the "Issue Summary" sheet
    issue_summary_sheet = workbook["Issue Summary"]
    summary_values = [row for row in issue_summary_sheet.iter_rows(values_only=True)][
        1:
    ]
    summary_values = [row for row in summary_values if any(row)]

    assert len(summary_values) == 2
    core_ids = set(row[1] for row in summary_values if row[1] is not None)
    assert core_ids == {"CDISC.SDTMIG.CG0999"}

    summary_by_dataset = {row[0]: row for row in summary_values if row[0] is not None}
    assert set(summary_by_dataset.keys()) == {"DM", "SUPPEC"}

    assert (
        summary_by_dataset["DM"][2]
        == "Issue with codelist definition in the Define-XML document."
    )
    assert (
        summary_by_dataset["SUPPEC"][2]
        == "rule evaluation error - evaluation dataset failed to build"
    )

    # Delete the excel file
    if os.path.exists(excel_file_path):
        os.remove(excel_file_path)
