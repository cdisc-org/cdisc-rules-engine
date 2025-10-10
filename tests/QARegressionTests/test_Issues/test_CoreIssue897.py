import os
import subprocess
import unittest
import openpyxl
import pytest
from conftest import get_python_executable
from QARegressionTests.globals import (
    issue_datails_sheet,
    rules_report_sheet,
    issue_sheet_record_column,
    issue_sheet_variable_column,
    issue_sheet_values_column,
)


@pytest.mark.regression
class TestGetXHTMLErrors(unittest.TestCase):
    def test_positive_dataset(self):
        # Run the command in the terminal
        command = [
            f"{get_python_executable()}",
            "-m",
            "core",
            "validate",
            "-s",
            "usdm",
            "-v",
            "4-0",
            "-dp",
            os.path.join(
                "tests",
                "resources",
                "CoreIssue897",
                "Positive_datasets.json",
            ),
            "-lr",
            os.path.join("tests", "resources", "CoreIssue897", "Rule.yml"),
        ]
        subprocess.run(command, check=True)

        # Get the latest created Excel file
        files = os.listdir()
        excel_files = [
            file
            for file in files
            if file.startswith("CORE-Report-") and file.endswith(".xlsx")
        ]
        excel_file_path = sorted(excel_files)[-1]
        # # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file_path)

        # Go to the "Issue Details" sheet
        sheet = workbook[issue_datails_sheet]

        record_column = sheet[issue_sheet_record_column]
        variables_column = sheet[issue_sheet_variable_column]
        values_column = sheet[issue_sheet_values_column]

        record_values = [cell.value for cell in record_column[1:]]
        variables_values = [cell.value for cell in variables_column[1:]]
        values_column_values = [cell.value for cell in values_column[1:]]

        # Remove None values using list comprehension
        record_values = [value for value in record_values if value is not None]
        variables_values = [value for value in variables_values if value is not None]
        values_column_values = [
            value for value in values_column_values if value is not None
        ]
        rules_values = [
            row for row in workbook[rules_report_sheet].iter_rows(values_only=True)
        ][1:]
        rules_values = [row for row in rules_values if any(row)]
        # Perform the assertion
        # Ensure only two negative values are caught
        assert rules_values[0][0] == "CORE-000409"
        assert rules_values[0][-1] == "SUCCESS"
        assert len(record_values) == 0
        assert len(variables_values) == 0
        assert len(values_column_values) == 0
        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)

    def test_negaive_dataset(self):
        # Run the command in the terminal
        command = [
            f"{get_python_executable()}",
            "-m",
            "core",
            "validate",
            "-s",
            "usdm",
            "-v",
            "4-0",
            "-dp",
            os.path.join(
                "tests",
                "resources",
                "CoreIssue897",
                "Negative_datasets.json",
            ),
            "-lr",
            os.path.join("tests", "resources", "CoreIssue897", "Rule.yml"),
        ]
        subprocess.run(command, check=True)

        # Get the latest created Excel file
        files = os.listdir()
        excel_files = [
            file
            for file in files
            if file.startswith("CORE-Report-") and file.endswith(".xlsx")
        ]
        excel_file_path = sorted(excel_files)[-1]
        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file_path)

        # --- Dataset Details ---
        dataset_sheet = workbook["Dataset Details"]
        dataset_values = [row for row in dataset_sheet.iter_rows(values_only=True)][1:]
        dataset_values = [row for row in dataset_values if any(row)]
        assert len(dataset_values) > 0
        assert dataset_values[0][0] == "NarrativeContentItem.xpt"
        assert dataset_values[0][-1] == 170

        # --- Issue Summary ---
        issue_summary_sheet = workbook["Issue Summary"]
        summary_values = [
            row for row in issue_summary_sheet.iter_rows(values_only=True)
        ][1:]
        summary_values = [row for row in summary_values if any(row)]
        assert len(summary_values) > 0
        assert summary_values[0][1] == "CORE-000409"
        assert summary_values[0][3] == 2

        # --- Issue Details ---
        issue_details_sheet = workbook["Issue Details"]
        details_values = [
            row for row in issue_details_sheet.iter_rows(values_only=True)
        ][1:]
        details_values = [row for row in details_values if any(row)]
        assert all(row[0] == "CORE-000409" for row in details_values)
        assert len(details_values) == 2

        # --- Rules Report ---
        rules_values = [
            row for row in workbook["Rules Report"].iter_rows(values_only=True)
        ][1:]
        rules_values = [row for row in rules_values if any(row)]
        assert len(rules_values) > 0
        assert rules_values[0][0] == "CORE-000409"
        assert rules_values[0][-1] == "SUCCESS"

        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)


# if __name__ == "__main__":
#     unittest.main()
