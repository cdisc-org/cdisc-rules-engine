import os
import subprocess
import openpyxl
import pytest
from conftest import get_python_executable
from QARegressionTests.globals import (
    dataset_details_sheet,
    issue_datails_sheet,
    rules_report_sheet,
    issue_sheet_variable_column,
    issue_sheet_coreid_column,
)


@pytest.mark.regression
def test_vlm_fallback_codelist_check():
    """
    Test for GitHub Issue #1443: Rule blocked: CDISC.SENDIG.49
    Validates that rules can use VLM (Value Level Metadata) columns as fallback
    when variable-level codelist is not available.
    
    Scenario:
    - VSORRESU variable has NO variable-level CodeListRef (empty ccode)
    - VSORRESU has VLM items with CodeListRef that match library standard
    - Rule should detect this mismatch and report ISSUE using VLM fallback columns
    """
    command = [
        f"{get_python_executable()}",
        "-m",
        "core",
        "validate",
        "-s",
        "sendig",
        "-v",
        "3-1",
        "-dp",
        os.path.join(
            "tests",
            "resources",
            "CoreIssue1443",
            "Dataset.json",
        ),
        "-lr",
        os.path.join("tests", "resources", "CoreIssue1443", "Rule.yml"),
        "-dxp",
        os.path.join("tests", "resources", "CoreIssue1443", "Define.xml"),
    ]
    subprocess.run(command, check=True)

    # Get the latest created Excel file
    files = os.listdir()
    excel_files = [
        file
        for file in files
        if file.startswith("CORE-Report-") and file.endswith(".xlsx")
    ]
    excel_file_path = sorted(excel_files)[-1]
    
    # Open the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)

    # Go to the "Issue Details" sheet
    sheet = workbook[issue_datails_sheet]

    # Check Variable(s) column
    variables_names_column = sheet[issue_sheet_variable_column]
    variables_names_values = [
        cell.value for cell in variables_names_column[1:] if cell.value is not None
    ]
    
    # DEBUG: print all issue details rows
    print("\n=== Issue Details (first 10 rows) ===")
    for row in sheet.iter_rows(min_row=1, max_row=11, values_only=True):
        if any(row):
            print(row)
    print(f"\nColumn I values: {variables_names_values}")
    
    # Verify that VSORRESU issue is detected
    assert len(variables_names_values) >= 1, "Expected at least one variable issue"
    assert any("VSORRESU" in str(val) for val in variables_names_values), \
        "Expected VSORRESU to be in issue variables"

    # Check Core ID
    core_id_column = sheet[issue_sheet_coreid_column]
    core_id_column_values = [
        cell.value for cell in core_id_column[1:] if cell.value is not None
    ]
    assert any("SEND49" in str(val) or "CDISC.SENDIG.49" in str(val) for val in core_id_column_values), \
        f"Expected SEND49 rule to report issues. Found: {core_id_column_values}"

    # Go to the "Rules Report" sheet
    rules_values = [
        row for row in workbook[rules_report_sheet].iter_rows(values_only=True)
    ][1:]
    rules_values = [row for row in rules_values if any(row)]
    
    # Verify rule execution
    assert len(rules_values) > 0, "Expected rule results in Rules Report"
    rule_ids = [row[0] for row in rules_values if row]
    assert any("SEND49" in str(rid) or "CDISC.SENDIG.49" in str(rid) for rid in rule_ids), \
        f"Expected SEND49 rule in Rules Report. Found: {rule_ids}"
    
    # Verify rule reported an issue
    for row in rules_values:
        if row and ("SEND49" in str(row[0]) or "CDISC.SENDIG.49" in str(row[0])):
            assert "ISSUE REPORTED" in str(row), \
                "Expected SEND49 to report an ISSUE"
            break

@pytest.mark.regression
def test_vlm_with_variable_level_codelist():
    """
    Test for GitHub Issue #1443 - Passing scenario
    Validates that rule does NOT flag VSORRESU when it HAS a variable-level codelist.
    """
    command = [
        f"{get_python_executable()}",
        "-m",
        "core",
        "validate",
        "-s",
        "sendig",
        "-v",
        "3-1",
        "-dp",
        os.path.join(
            "tests",
            "resources",
            "CoreIssue1443",
            "Dataset.json",
        ),
        "-lr",
        os.path.join("tests", "resources", "CoreIssue1443", "Rule.yml"),
        "-dxp",
        os.path.join("tests", "resources", "CoreIssue1443", "Define_with_codelist.xml"),
    ]
    subprocess.run(command, check=True)

    # Get the latest created Excel file
    files = os.listdir()
    excel_files = [
        file
        for file in files
        if file.startswith("CORE-Report-") and file.endswith(".xlsx")
    ]
    excel_file_path = sorted(excel_files)[-1]
    
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[issue_datails_sheet]

    # Check Variable(s) column
    variables_names_column = sheet[issue_sheet_variable_column]
    variables_names_values = [
        cell.value for cell in variables_names_column[1:] if cell.value is not None
    ]
    
    # Verify that VSORRESU is NOT flagged when variable-level codelist is present
    assert not any("VSORRESU" in str(val) for val in variables_names_values), \
        "Expected VSORRESU NOT to be flagged when variable-level codelist is present"