import os
import subprocess
import unittest

import openpyxl
import pytest
from conftest import get_python_executable
from QARegressionTests.globals import (
    issue_datails_sheet,
    rules_report_sheet,
)


@pytest.mark.regression
class TestCoreIssue720(unittest.TestCase):
    def test_negative_dataset(self):
        """Negative scenario: SPECIES missing -> expect one populated issue."""
        command = [
            f"{get_python_executable()}",
            "-m",
            "core",
            "validate",
            "-s",
            "usdm",
            "-v",
            "4-0",
            "-dp",
            os.path.join("tests", "resources", "CoreIssue720", "Invalid_datasets.json"),
            "-lr",
            os.path.join("tests", "resources", "CoreIssue720", "Rule.yml"),
        ]
        subprocess.run(command, check=True)

        excel_files = [
            f
            for f in os.listdir()
            if f.startswith("CORE-Report-") and f.endswith(".xlsx")
        ]
        assert excel_files, "No CORE report generated"
        excel_file_path = sorted(excel_files)[-1]
        wb = openpyxl.load_workbook(excel_file_path)

        # Conformance Details
        assert "Conformance Details" in wb.sheetnames
        conf_rows = [
            row for row in wb["Conformance Details"].iter_rows(values_only=True)
        ]
        assert conf_rows[6][0] == "Standard"
        assert conf_rows[6][1] == "USDM"
        assert conf_rows[7][0] == "Version"
        assert conf_rows[7][1] == "V4.0"

        # Entity Details
        entity_rows = [row for row in wb["Entity Details"].iter_rows(values_only=True)][
            1:
        ]
        entity_rows = [r for r in entity_rows if any(r)]
        assert len(entity_rows) >= 2
        assert entity_rows[0][0] == "DM" and entity_rows[0][1] == 4
        assert entity_rows[1][0] == "TS" and entity_rows[1][1] == 4

        # Issue Summary
        summary_rows = [row for row in wb["Issue Summary"].iter_rows(values_only=True)][
            1:
        ]
        summary_rows = [r for r in summary_rows if any(r)]
        assert summary_rows[0][0] == "dm.xpt"
        assert summary_rows[0][1] == "CDISC.SENDIG.105"
        assert summary_rows[0][3] == 1

        # Issue Details
        details_rows = [
            row for row in wb[issue_datails_sheet].iter_rows(values_only=True)
        ][1:]
        details_rows = [r for r in details_rows if any(r)]
        assert details_rows, "Issue Details should have at least one populated row"
        first_issue = details_rows[0]
        assert first_issue[0] == "CDISC.SENDIG.105"
        assert first_issue[1] == "SEND105"
        assert "SPECIES in not present" in first_issue[2]
        assert first_issue[3] == "fully executable"
        assert first_issue[4] == "dm.xpt"
        assert first_issue[7] == "$distinct_tsparmcd, $distinct_txparmcd, SPECIES"
        assert "Not in dataset" in first_issue[8]

        # Rules Report
        rules_rows = [
            row for row in wb[rules_report_sheet].iter_rows(values_only=True)
        ][1:]
        rules_rows = [r for r in rules_rows if any(r)]
        assert rules_rows, "Rules Report must have at least one populated row"
        assert rules_rows[0][0] == "CDISC.SENDIG.105"
        assert rules_rows[0][-1] == "SUCCESS"

        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)

    def test_positive_dataset(self):
        """Positive scenario: SPECIES present -> expect no issue rows."""
        command = [
            f"{get_python_executable()}",
            "-m",
            "core",
            "validate",
            "-s",
            "usdm",
            "-v",
            "4-0",
            "-dp",
            os.path.join("tests", "resources", "CoreIssue720", "Valid_datasets.json"),
            "-lr",
            os.path.join("tests", "resources", "CoreIssue720", "Rule.yml"),
        ]
        subprocess.run(command, check=True)

        excel_files = [
            f
            for f in os.listdir()
            if f.startswith("CORE-Report-") and f.endswith(".xlsx")
        ]
        assert excel_files, "No CORE report generated"
        excel_file_path = sorted(excel_files)[-1]
        wb = openpyxl.load_workbook(excel_file_path)

        # Issue Summary empty
        summary_rows = [row for row in wb["Issue Summary"].iter_rows(values_only=True)][
            1:
        ]
        summary_rows = [r for r in summary_rows if any(r)]
        assert summary_rows == []

        # Issue Details empty
        details_rows = [
            row for row in wb[issue_datails_sheet].iter_rows(values_only=True)
        ][1:]
        details_rows = [r for r in details_rows if any(r)]
        assert details_rows == []

        # Rules Report has success row
        rules_rows = [
            row for row in wb[rules_report_sheet].iter_rows(values_only=True)
        ][1:]
        rules_rows = [r for r in rules_rows if any(r)]
        assert len(rules_rows) == 1
        assert rules_rows[0][0] == "CDISC.SENDIG.105"
        assert rules_rows[0][-1] == "SUCCESS"

        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)
