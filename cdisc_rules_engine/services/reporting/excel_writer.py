from io import BytesIO
from os import unlink
from tempfile import NamedTemporaryFile

import openpyxl
from openpyxl.styles import Alignment


def excel_workbook_to_stream(wb):
    stream = None
    tmp = NamedTemporaryFile(
        delete=False
    )  # create tmp file, we'll controll when to delete it
    # close, NamedTemporaryFile() already opended it, we need it closed for wb.save()
    # or we'll get permission denied errors
    tmp.close()
    wb.save(tmp.name)  # save workbook to tmp file
    with open(tmp.name, "rb") as f:
        stream = f.read()  # open for reading
    unlink(tmp.name)  # delete
    return stream


def excel_update_worksheet(ws, rows, align_params=None, fill_empty_rows=False):
    for row_num, row_data in enumerate(rows, 2):
        for col_num, col_data in enumerate(row_data, 1):
            if fill_empty_rows and (row_data[1] == "" or row_data[1] is None):
                # Codelist is empty for Code Rows. Change background color
                ws.cell(row=row_num, column=col_num).value = col_data
                ws.cell(row=row_num, column=col_num).fill = openpyxl.styles.PatternFill(
                    start_color="ccffff", end_color="ccffff", fill_type="solid"
                )
                ws.cell(row=row_num, column=col_num).alignment = Alignment(
                    **align_params
                )
            else:
                ws.cell(row=row_num, column=col_num).value = col_data
                ws.cell(row=row_num, column=col_num).alignment = Alignment(
                    **align_params
                )


def excel_open_workbook(template_buffer):
    byte_stream = BytesIO(template_buffer)
    return openpyxl.load_workbook(byte_stream)
