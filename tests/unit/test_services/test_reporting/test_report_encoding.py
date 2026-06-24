import os
import json
import openpyxl
from unittest.mock import MagicMock
from io import BytesIO
from cdisc_rules_engine.services.reporting.json_report import JsonReport
from cdisc_rules_engine.services.reporting.excel_report import ExcelReport
from cdisc_rules_engine.services.reporting.report_metadata_item import (
    ReportMetadataItem,
)


def test_json_report_encoding(tmp_path):
    # Setup
    mock_args = MagicMock()
    mock_args.output = str(tmp_path / "test_report_utf8")
    mock_args.output_format = ["JSON"]
    mock_args.raw_report = False
    mock_args.encoding = "utf-8"

    mock_report_data = MagicMock()
    mock_report_data.data_sheets = {
        "Test Sheet": {"Message": "Hello World", "Unicode": "こんにちは世界"}
    }

    report = JsonReport(mock_report_data, mock_args)
    report.write_report()

    output_file = f"{mock_args.output}.json"
    assert os.path.exists(output_file)

    with open(output_file, "r", encoding="utf-8") as f:
        content = json.load(f)
        assert content["Test_Sheet"]["Unicode"] == "こんにちは世界"


def test_excel_report_unicode_support(tmp_path):
    # Setup
    mock_args = MagicMock()
    mock_args.output = str(tmp_path / "test_report_excel")
    # Show report folder
    print(f"tmp_path contents: {os.listdir(tmp_path)}")
    mock_args.output_format = ["XLSX"]
    mock_args.max_report_rows = None
    # Encoding arg shouldn't matter for Excel output content as it is binary/xml
    mock_args.encoding = "utf-8"

    mock_report_data = MagicMock()

    # Create a simple empty excel file in memory as template
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    wb.create_sheet("Conformance Details")

    template_stream = BytesIO()
    wb.save(template_stream)
    template_stream.seek(0)

    mock_report_data.data_sheets = {
        "Test Sheet": [
            ReportMetadataItem(name="Test Item", row=1, value="こんにちは世界")
        ],
        "Conformance Details": [],
    }

    report = ExcelReport(mock_report_data, mock_args, template_stream)
    report.write_report()

    output_file = f"{mock_args.output}.xlsx"
    assert os.path.exists(output_file)

    wb_out = openpyxl.load_workbook(output_file)
    ws_out = wb_out["Test Sheet"]
    # ExcelReport writes ReportMetadataItem value to column 2
    assert ws_out.cell(row=1, column=2).value == "こんにちは世界"
